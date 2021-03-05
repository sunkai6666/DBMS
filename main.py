
import os
import sys
import pandas as pd
import tableop
import createtable
import openpyxl
import re
import analysis
import tableop
import userlogin
import start
i=0
path=sys.path[0]
basename=' '
flag=0
flag1=0
user=[]
while True :
    if flag1==0:
        user=userlogin.login()
        userlogin.check(user[0],user[1])
        flag1=1
    name=input()
    name=name.lower()
    def token(name):
        name=name.rstrip(';')
        name=name.replace(',',' ')
        inputlist=name.split(' ')
        return inputlist
    input_list=token(name)
    if (input_list[0] )== 'use':
        current_base=openpyxl.load_workbook(path+'/base/'+"basename.xlsx")
        current_table=current_base.active
        for column in current_table['A']:
            if (input_list[1]==column.value):
                basename=input_list[1]
                flag=1
        if flag==0:
                print("不存在此数据库")
    elif(input_list[0]=='help'):
        current_base=openpyxl.load_workbook(path+'/base/'+basename+"/tableinformation.xlsx")
        if(input_list[1]=='database'):
            print(current_base.sheetnames)
        elif(input_list[1]=='table'):
            if(input_list[2] not in current_base.sheetnames):
                print('此表不存在')
            else:
                data=pd.read_excel(path+'/base/'+basename+'/'+"tableinformation.xlsx",sheet_name='help'+input_list[2])
                print(data)              
    elif (input_list[0]=='create'):
        if user[0]=='sk':
            if input_list[1]=="database":
                start.start(path,input_list[2])
            else:    
                createtable.create(name,basename)
    elif (input_list[0]=='insert'): 
        tableop.insert(name,basename)
    elif (input_list[0]=='update'):
        tableop.update(name,basename)
    elif(input_list[0]=='select'):
        analysis.selectop(input_list,basename,name)
    elif(input_list[0]==('delete')):
        analysis.delete(name,basename,input_list)