import os
import sys
import openpyxl
import pandas as pd
import string
import re
import xlrd
def insert (name,basename):
    path=sys.path[0]
    dbpath=path+'/base/'+basename+'/'
    current_base=openpyxl.load_workbook(dbpath+"tableinformation.xlsx")
    if 'values' in name:
        result=re.search('insert into (.*) \((.*)\) values\((.*)\);$',name)
        table = current_base[result.group(1)]
        columns=result.group(2).split(',')
        values=result.group(3).split(',')
        k_v=dict(zip(columns,values))
        k=list(k_v.keys())
        v=list(k_v.values())
        k_length=len(k)
        table_rows = table.max_row
        table_columns = table.max_column
        for i in range(0,k_length) : 
            for j in range(1,table_columns+1):
               if(k[i]==table.cell(1,j).value):
                   table.cell(table_rows+1,j).value=v[i]
        chk_len = len(values)
        if (chk_len>table_columns):
            print ('插入失败')
        else :
            #table.append()
            pass
        current_base.save(dbpath+"tableinformation.xlsx")
def update(name,basename):
    path=sys.path[0]
    dbpath=path+'/base/'+basename+'/'
    current_base=openpyxl.load_workbook(dbpath+"tableinformation.xlsx")
    if 'where' not in name:    
        result=re.search('update (.*) set (.*);$',name)
        table = current_base[result.group(1)]
        setlist=result.group(2)
        table_rows = table.max_row
        table_columns = table.max_column
        list=['+','-','*','/']
        for z in range(0,4):
            if str(list[z]) in setlist:
                setlist=setlist.replace('=',' ')
                setlist=setlist.replace(list[z],' ')
                setlist=setlist.split(' ')
                for i in range(1,table_columns+1):
                    if(table.cell(1,i).value==setlist[0]):
                        for j in range(2,table_rows+1):
                            s=table.cell(j,i).value
                            if list[z]=='+':
                                table.cell(j,i).value=int(s)+int(setlist[2])
                            elif list[z]=='-':
                                table.cell(j,i).value=int(s)-int(setlist[2])
                            elif list[z]=='*':
                                table.cell(j,i).value=int(s)*int(setlist[2])
                            else :
                                table.cell(j,i).value=int(s)/int(setlist[2])
        current_base.save(dbpath+"tableinformation.xlsx")
    elif 'where'  in name:
        result=re.search('update (.*) set (.*) where (.*);$',name)
        table = current_base[result.group(1)]
        set_list=result.group(2).replace('=',' ')
        set_list=set_list.split(' ')
        v_list=result.group(3).replace('=',' ')
        v_list=v_list.split(' ')
        table_rows = table.max_row
        table_columns = table.max_column
        r=0
        flag=0
        for i in range(1,table_columns+1):
            if(table.cell(1,i).value==v_list[0]):
                for j in range(2,table_rows+1):
                    if(table.cell(j,i).value==v_list[1]):
                        r=j
                        flag=1
                        break
            if(flag==1):
                break
        for i in range(1,table_columns+1):
            if(table.cell(1,i).value==set_list[0] ):
                print('2')
                table.cell(r,i).value=set_list[1]
        current_base.save(dbpath+"tableinformation.xlsx")
    elif ():
        print("不存在该属性")
def select (name,basename):
    path=sys.path[0]
    dbpath=path+'/base/'+basename+'/'
    xlsx=pd.ExcelFile(dbpath+"tableinformation.xlsx")
    data=pd.read_excel(xlsx,sheet_name='wang')

#   print(sqldf(name,locals()))
    #print(format(data))
