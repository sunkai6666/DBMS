import pandas as pd
import sys
import re
import openpyxl
def selectop(token,basename,name):
    path=sys.path[0]
    dbpath=path+'/base/'+basename+'/'
    from_index=token.index("from")
    column=token[1:from_index]
    data=pd.read_excel(dbpath+"tableinformation.xlsx",sheet_name=token[from_index+1])
    if "where" not in token: 
        if "*" in token:
            print(data)
        else:
            print(data.loc[:,column])
    else :
        name=re.search("() where (.*);",name)
        name=name.group(2)
        if 'in' in name :
            name=name.replace("(","[").replace(")","]")
            data=data.query(name)
        elif 'between' in name and 'not' not in name:
            name=name.split(' ')
            sql=name[0]+">="+name[2]+' and '+name[0]+"<="+name[4]
            data=data.query(sql)
        elif 'between' in name and 'not' in name:
            name=name.split(' ')
            sql=name[0]+"<"+name[3]+' or '+name[0]+">"+name[5]
            data=data.query(sql)
        else:
            name=name.replace("<>","!=")
            name=name.replace("=","==")
            name=name.replace(">==",">=")
            name=name.replace("<==","<=")
            name=name.replace("!==","!=")
            data=data.query(name)
        if '*' not in column:
            print(data.loc[:,column])
        else:
            print(data)
def delete (name,basename,token):
    path=sys.path[0]
    dbpath=path+'/base/'+basename+'/'
    name=re.search("() where (.*);",name)
    current_base=openpyxl.load_workbook(dbpath+"tableinformation.xlsx")
    table = current_base[token[2]]
    table_columns = table.max_column
    table_rows = table.max_row
    column_list=[]
    if "where" not in token: 
        table.delete_rows(2,table_rows-1)
    else:
        name=name.group(2).replace('=',' ').replace("'",'')
        name=name.split(' ')
        for i in range(1,table_columns+1):
            if(table.cell(1,i).value==name[0]):
                for j in range(2,table_rows+1):
                    if(str(table.cell(j,i).value)==name[1]):
                        table.delete_rows(j,1)
    current_base.save(dbpath+"tableinformation.xlsx")

