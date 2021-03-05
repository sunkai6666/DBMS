import os
import sys
import openpyxl
import pandas as pa
import re
def create(name,basename):
    path=sys.path[0]
    dbpath=path+'/base/'+basename+'/'
    result=re.search('create table (.*) \((.*)\);$',name)
    table_name=result.group(1)
    if (not os.path.exists(dbpath+"tableinformation.xlsx")):
       current_base=openpyxl.Workbook()
       current_base.save(dbpath+"tableinformation.xlsx") 
       current_base.close()
    current_base=openpyxl.load_workbook(dbpath+"tableinformation.xlsx")
    #print(current_base.sheetnames)
    if (table_name not in current_base.sheetnames):
       tablelens=len(current_base.sheetnames)
       #print(tablelens)
       #attribute_list=attribute_list.replace(',',' ')
       #attribute_list=attribute_list.split(' ')
       columns_list = re.findall('\((.*)\)',name)[0].split(',')
       length = len(columns_list)
       column_names = []
       list=['Filed','Type','Null','Key','Default','Extra']
       helptable=current_base.create_sheet('help'+table_name)
       helptable.append(list)
       for i in range(length):
         column = columns_list[i].split(' ')
         helptable.cell(2+i,1).value=column[0]
         helptable.cell(2+i,2).value=column[1]
         for key in column[2:]:
            if(key=='primary'):
              helptable.cell(2+i,4).value='PRY'
            elif(key=='not'):
               helptable.cell(2+i,3).value='NO'
            elif(key=='unique'):
               helptable.cell(2+i,6).value='UNIQUE'
         #for key in column:
         #   pass
         column_names.append(column[0])
       #current_base['Sheet'].append(attribute_list)
       table=current_base.create_sheet(table_name)
       #print(current_base.sheetnames)
       table.append(column_names)
       current_base.save(dbpath+"tableinformation.xlsx")
       current_base.close()
    else :
        print ("表已存在")