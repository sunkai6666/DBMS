import os
import sys
import hashlib
import openpyxl
def login ():
    if not os.path.exists(sys.path[0]+'/user.xlsx'):
        current_base=openpyxl.Workbook()
        current_base.save(sys.path[0]+'/user.xlsx') 
        current_base.close()
    print("user :")
    user=input()
    print("password :")
    password=input()
    md5_pw=hashlib.md5(password.encode("utf-8")).hexdigest()
    userlist=[user,md5_pw]
    #check(user,md5_pw)
    return userlist
def check(user,password):
    flag2=1
    current_base=openpyxl.load_workbook(sys.path[0]+'/user.xlsx')
    table=current_base.active
    table_rows = table.max_row
    for i in range(1,table_rows+1):
        if (user==table.cell(i,1).value):
            flag2=0
            if(password==table.cell(i,2).value):
                print(user+"欢迎登录")
            else:
                print("密码错误")
                login()
    if flag2==1:
        print("不存在该用户，您可以使用 create user 创建")
        newuser=input()
        new=login()
        table.append(new)
        current_base.save(sys.path[0]+'/user.xlsx') 
    